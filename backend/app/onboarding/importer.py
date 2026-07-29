"""文件导入逻辑 — 从 MVP 迁移并增强"""

import re
from datetime import date
from pathlib import Path

import openpyxl

from ..core.logging import get_logger

logger = get_logger(__name__)


def import_excel(filepath: str, db, user: dict) -> dict:
    """
    导入 Excel 文件:
    1. 自动检测表头和列类型
    2. 去重检测 (table_name + data_period)
    3. 建表 + 导入数据
    4. 注册到 data_snapshots
    5. 自动生成指标
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_result = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # 检测表头位置
        header_row_idx = _detect_header(rows)
        if header_row_idx is None:
            sheets_result.append({"sheet": sheet_name, "status": "skipped", "reason": "未检测到表头"})
            continue

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_row_idx])]
        data_rows = rows[header_row_idx + 1:]

        if not data_rows:
            sheets_result.append({"sheet": sheet_name, "status": "skipped", "reason": "无数据行"})
            continue

        # 检测: 指标需求清单 (含"指标名称"列) → 导入到 metric_registry
        if any("指标名称" in h for h in headers):
            count = _import_metric_requirements(db, sheet_name, headers, data_rows)
            sheets_result.append({
                "sheet": sheet_name, "status": "imported",
                "metrics_imported": count, "type": "metric_requirements",
            })
            continue

        # 生成表名 (含文件名上下文避免冲突)
        base_name = _safe_table_name(sheet_name)
        # 如果sheet名是纯日期格式，用文件名+sheet名
        if re.match(r'^\d{4}[-_]\d{2}$', str(sheet_name)):
            import os
            file_stem = os.path.splitext(os.path.basename(filepath))[0]
            file_prefix = _safe_table_name(file_stem)[:30]
            table_name = f"{file_prefix}_{base_name}"[:64]
        else:
            table_name = base_name

        # 推断列类型
        col_types = _infer_column_types(headers, data_rows)

        # 推断数据期间
        data_period = _infer_period(data_rows, col_types)

        # 去重检测
        existing = db.execute_one(
            "SELECT snapshot_id FROM data_snapshots WHERE table_name = ? AND data_period = ?",
            (table_name, data_period),
        )
        if existing:
            sheets_result.append({
                "sheet": sheet_name, "table_name": table_name, "status": "skipped",
                "reason": f"数据已存在 (snapshot #{existing['snapshot_id']})",
            })
            continue

        # 建表
        _create_table(db, table_name, headers, col_types)

        # 导入数据
        count = _import_data(db, table_name, headers, data_rows, col_types)

        # 注册快照
        snapshot_id = db.insert_and_get_id(
            "INSERT INTO data_snapshots (table_name, data_period, ingestion_time, description, total_rows, uploaded_by) "
            "VALUES (?, ?, datetime('now'), ?, ?, ?)",
            (table_name, data_period, f"{sheet_name} - {data_period}", count, user["user_id"]),
        )

        # 更新 snapshot_id
        db.execute_write(
            f"UPDATE \"{table_name}\" SET snapshot_id = ? WHERE snapshot_id = 0",
            (snapshot_id,),
        )

        # 自动生成指标
        metrics_count = 0
        try:
            metrics_count = _auto_generate_metrics(db, table_name, headers, col_types)
        except Exception as e:
            logger.warning("指标生成失败: %s", e)

        # 自动注册KB同义词 (仅第一次导入时)
        if metrics_count > 0:
            try:
                _auto_register_kb_synonyms(headers, col_types)
            except Exception as e:
                logger.warning("KB同义词注册失败: %s", e)

        sheets_result.append({
            "sheet": sheet_name, "table_name": table_name, "status": "imported",
            "data_period": data_period, "row_count": count, "columns": len(headers),
            "snapshot_id": snapshot_id,
            "metrics_generated": metrics_count,
        })

    wb.close()

    # 自动连接 pending 指标到数据表
    try:
        _auto_connect_pending_metrics(db)
    except Exception:
        pass

    # 刷新指标加载器
    try:
        from ...semantic.loader import MetricLoader
        MetricLoader(db).reload()
    except Exception:
        pass

    return {
        "type": "import_result",
        "sheets": sheets_result,
        "total_imported": sum(1 for s in sheets_result if s["status"] == "imported"),
        "total_skipped": sum(1 for s in sheets_result if s["status"] == "skipped"),
    }


def _detect_header(rows: list) -> int | None:
    """检测表头行: 前10行中第一个含2+非空单元格的行"""
    for i in range(min(10, len(rows))):
        non_empty = sum(1 for v in rows[i] if v is not None and str(v).strip())
        if non_empty >= 2:
            return i
    return None


def _safe_table_name(name: str) -> str:
    """去除特殊字符, 转蛇形命名"""
    name = re.sub(r'[^\w一-鿿]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    return name or "imported_table"


def _infer_column_types(headers: list, data_rows: list) -> dict:
    """推断列类型: DATE / REAL / INTEGER / TEXT"""
    types = {}
    for i, h in enumerate(headers):
        samples = [row[i] for row in data_rows[:20] if i < len(row) and row[i] is not None]
        if not samples:
            types[h] = "TEXT"
            continue

        # 尝试将字符串转为数字
        def try_num(v):
            if isinstance(v, (int, float)): return v
            if isinstance(v, str):
                try: return int(v)
                except: pass
                try: return float(v)
                except: pass
            return v

        # 检测日期
        date_count = sum(1 for v in samples if isinstance(v, (date,)) or
                         (isinstance(v, str) and re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', str(v))))
        if date_count > len(samples) * 0.5:
            types[h] = "DATE"
            continue

        # 检测数字 (含字符串数字)
        nums = [try_num(v) for v in samples]
        num_count = sum(1 for v in nums if isinstance(v, (int, float)))
        if num_count == len(samples):
            if all(isinstance(v, int) for v in nums):
                types[h] = "INTEGER"
            else:
                types[h] = "REAL"
            continue

        types[h] = "TEXT"

    return types


def _infer_period(data_rows: list, col_types: dict) -> str:
    """从日期列推断数据期间"""
    today = date.today()
    for h, t in col_types.items():
        if t == "DATE":
            dates = []
            for row in data_rows[:20]:
                val = row[list(col_types.keys()).index(h)] if h in col_types else None
                if val:
                    if isinstance(val, date):
                        dates.append(val)
                    elif isinstance(val, str):
                        m = re.match(r'(\d{4})-(\d{1,2})', val)
                        if m:
                            dates.append(date(int(m.group(1)), int(m.group(2)), 1))
            if dates:
                min_d = min(dates)
                max_d = max(dates)
                if min_d.strftime("%Y-%m") == max_d.strftime("%Y-%m"):
                    return min_d.strftime("%Y-%m")
                return f"{min_d.strftime('%Y-%m')}~{max_d.strftime('%Y-%m')}"
    return f"{today.year}-{today.month:02d}"


def _create_table(db, table_name: str, headers: list, col_types: dict):
    """创建表"""
    type_map = {"DATE": "TEXT", "REAL": "REAL", "INTEGER": "INTEGER", "TEXT": "TEXT"}
    cols = ['_row_id INTEGER PRIMARY KEY AUTOINCREMENT', 'snapshot_id INTEGER DEFAULT 0']
    for h in headers:
        safe_h = re.sub(r'[^\w一-鿿]', '_', str(h))
        cols.append(f'"{safe_h}" {type_map.get(col_types[h], "TEXT")}')

    # 如果表已存在, 删除重建
    try:
        db.execute_write(f'DROP TABLE IF EXISTS "{table_name}"')
    except Exception:
        pass

    db.execute_write(f'CREATE TABLE "{table_name}" ({", ".join(cols)})')


def _import_data(db, table_name: str, headers: list, data_rows: list, col_types: dict) -> int:
    """导入数据"""
    safe_headers = [re.sub(r'[^\w一-鿿]', '_', str(h)) for h in headers]
    placeholders = ','.join(['?'] * (len(headers) + 1))
    cols = ','.join(['snapshot_id'] + [f'"{h}"' for h in safe_headers])

    count = 0
    for row in data_rows:
        values = [0]  # snapshot_id = 0, 后续更新
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val is None:
                values.append(None)
            elif col_types[h] == "DATE":
                if isinstance(val, date):
                    values.append(val.strftime("%Y-%m-%d"))
                else:
                    values.append(str(val))
            else:
                values.append(val)

        try:
            db.execute_write(f'INSERT INTO "{table_name}" ({cols}) VALUES ({placeholders})', tuple(values))
            count += 1
        except Exception:
            pass

    return count


def _import_metric_requirements(db, sheet_name: str, headers: list, data_rows: list) -> int:
    """导入指标需求清单 → metric_registry (status=pending, 等待数据接入)"""
    count = 0
    # 找到各列索引
    name_idx = next((i for i, h in enumerate(headers) if "指标名称" in h), 0)
    explain_idx = next((i for i, h in enumerate(headers) if "解释" in h), None)
    formula_idx = next((i for i, h in enumerate(headers) if "公式" in h), None)
    source_idx = next((i for i, h in enumerate(headers) if "来源" in h), None)

    for row in data_rows:
        if len(row) <= name_idx:
            continue
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        if not name or len(name) < 2 or name in ("指标名称", "None"):
            continue

        explanation = str(row[explain_idx])[:500] if explain_idx and explain_idx < len(row) and row[explain_idx] else ""
        formula = str(row[formula_idx])[:200] if formula_idx and formula_idx < len(row) and row[formula_idx] else ""
        source = str(row[source_idx])[:200] if source_idx and source_idx < len(row) and row[source_idx] else ""

        # 注册指标 (已存在的补充解释/公式)
        try:
            exist = db.execute_one("SELECT metric_id, explanation FROM metric_registry WHERE name=?", (name,))
            if exist:
                if not exist.get("explanation") and explanation:
                    db.execute_write(
                        "UPDATE metric_registry SET explanation=?, formula=?, source=?, category=? WHERE metric_id=?",
                        (explanation, formula, source, sheet_name, exist["metric_id"]))
            else:
                db.execute_write(
                    "INSERT INTO metric_registry (name, category, status, complexity, explanation, formula, source) "
                    "VALUES (?, ?, 'pending', 'L1', ?, ?, ?)",
                    (name, sheet_name, explanation, formula, source))
            count += 1
        except Exception as e:
            logger.debug("指标注册跳过: %s - %s", name, e)

    # 自动注册KB同义词
    metrics_added = [str(row[name_idx]).strip() for row in data_rows
                     if len(row) > name_idx and row[name_idx] and str(row[name_idx]).strip()]
    if metrics_added:
        _register_requirement_synonyms(sheet_name, metrics_added)

    logger.info("指标需求导入: %s → %d 个指标", sheet_name, count)
    return count


def _register_requirement_synonyms(category: str, metric_names: list[str]):
    """为导入的指标需求自动注册KB同义词"""
    import yaml
    from pathlib import Path

    kb_path = Path(__file__).parent.parent.parent / "metrics" / "enterprise_kb.yaml"
    try:
        with open(kb_path, "r", encoding="utf-8") as f:
            kb = yaml.safe_load(f) or {}
    except Exception:
        kb = {}

    synonyms = kb.get("synonyms", {})
    if synonyms is None:
        synonyms = {}

    added = 0
    for name in metric_names:
        if name and name not in synonyms:
            synonyms[name] = name
            added += 1
        # 简化名（去掉前缀/后缀的常见词）
        short = name.replace("年度累计", "").replace("本期", "").replace("数量", "").replace("金额", "")
        if short and short != name and len(short) >= 2 and short not in synonyms:
            synonyms[short] = name
            added += 1

    if added > 0:
        kb["synonyms"] = synonyms
        with open(kb_path, "w", encoding="utf-8") as f:
            yaml.dump(kb, f, allow_unicode=True, default_flow_style=False)
        logger.info("已注册 %d 个指标需求同义词", added)


def _auto_connect_pending_metrics(db) -> int:
    """自动连接指标到数据表: 关键词匹配 → 生成SQL → 激活 (含已有指标更新)"""
    # 1. 更新已有指标的解释和SQL (来自需求清单导入但SQL可能过时)
    all_metrics = db.execute(
        "SELECT metric_id, name, explanation, formula, status FROM metric_registry"
    )
    if not all_metrics:
        return 0

    # 2. 处理所有指标: pending激活, available更新SQL指向新表
    pending = [m for m in all_metrics if m["status"] == "pending"]
    existing = [m for m in all_metrics if m["status"] == "available"]

    targets = pending + existing

    tables = db.execute("SELECT DISTINCT table_name FROM data_snapshots")
    table_names = [t["table_name"] for t in tables]

    # 关键词 → 表名映射
    KEYWORD_TABLE_MAP = {
        "中标": "中标管理表",
        "签约": "合同管理表",
        "合同": "合同管理表",
        "商机": "商机管理表",
        "客户": "合同管理表",
        "应收": "应收账款表",
        "逾期": "应收账款表",
        "回款": "应收账款表",
    }

    connected = 0
    for m in targets:
        name = m["name"]
        matched_table = None
        for kw, tbl in KEYWORD_TABLE_MAP.items():
            if kw in name:
                for tn in table_names:
                    if tn.startswith(tbl):
                        matched_table = tn
                        break
                if matched_table:
                    break
        if not matched_table:
            continue

        schema = db.get_table_schema(matched_table)
        numeric_cols = [c["name"] for c in schema
            if c["name"] not in ("_row_id", "snapshot_id")
            and any(t in c["type"].upper() for t in ("REAL", "INTEGER", "FLOAT"))]
        if not numeric_cols:
            continue

        col = numeric_cols[0]
        safe_col = col.replace('"', '""')

        # 聚合方式判断 (同前)
        if any(kw in name for kw in ("各地市", "各业务线", "各区域", "分布", "排名", "排行")):
            group_col = None
            if "各地市" in name or "城市" in name or "区域" in name:
                for c in schema:
                    if c["name"] in ("region", "所属区域", "区域") or "区域" in c["name"]:
                        group_col = c["name"]; break
            if "业务线" in name or "板块" in name:
                for c in schema:
                    if c["name"] in ("business_line", "业务线") or "业务线" in c["name"]:
                        group_col = c["name"]; break
            if group_col:
                safe_g = group_col.replace('"', '""')
                sql = f'SELECT "{safe_g}" AS label, ROUND(SUM("{safe_col}"),2) AS value FROM "{matched_table}" GROUP BY "{safe_g}" ORDER BY value DESC'
                fmt = "table"
            else:
                sql = f'SELECT ROUND(SUM("{safe_col}"),2) AS value FROM "{matched_table}"'
                fmt = "number"
        elif any(kw in name for kw in ("数量", "个数", "项目数", "客户数", "总数")):
            sql = f'SELECT COUNT(*) AS value FROM "{matched_table}"'
            fmt = "integer"
        elif any(kw in name for kw in ("率", "比", "转化")):
            sql = f'SELECT ROUND(AVG("{safe_col}"),2) AS value FROM "{matched_table}"'
            fmt = "percent"
        else:
            sql = f'SELECT ROUND(SUM("{safe_col}"),2) AS value FROM "{matched_table}"'
            fmt = "number"

        # 更新: pending→available, 已有→更新SQL+explanation
        if m["status"] == "pending":
            db.execute_write(
                "UPDATE metric_registry SET status='available', table_name=?, sql_template=?, result_format=?, complexity='L1' WHERE metric_id=?",
                (matched_table, sql, fmt, m["metric_id"]),
            )
        else:
            # 已有指标: 更新SQL指向新数据表
            db.execute_write(
                "UPDATE metric_registry SET table_name=?, sql_template=?, result_format=? WHERE metric_id=?",
                (matched_table, sql, fmt, m["metric_id"]),
            )
        connected += 1

    logger.info("自动连接: %d 个指标已更新/激活", connected)
    return connected


def _humanize_metric_name(col_name: str, agg: str) -> str:
    """生成人性化指标名: '合同金额(万元)' + '合计' → '合同金额 合计'"""
    # 去掉括号中的单位: 合同金额(万元) → 合同金额
    clean = re.sub(r'\([^)]*\)', '', col_name).strip()
    # 去掉下划线残留
    clean = re.sub(r'_+', ' ', clean).strip()
    if not clean:
        clean = col_name
    return f"{clean} {agg}"


def _auto_generate_metrics(db, table_name: str, headers: list, col_types: dict) -> int:
    """自动生成指标: 每个数值列生成合计/平均, 人性化命名"""
    count = 0
    safe_headers = [re.sub(r'[^\w一-鿿]', '_', str(h)) for h in headers]

    for i, h in enumerate(headers):
        safe_h = safe_headers[i]
        if col_types[h] in ("REAL", "INTEGER"):
            clean_name = _humanize_metric_name(h, "").strip()

            # 合计 (不同表同名指标自动跳过)
            metric_name = f"{clean_name} 合计"
            sql_sum = f'SELECT ROUND(SUM("{safe_h}"), 2) AS value FROM "{table_name}"'
            if not db.execute_one("SELECT 1 FROM metric_registry WHERE name=?", (metric_name,)):
                try:
                    db.execute_write(
                        "INSERT INTO metric_registry (name, display_name, category, status, complexity, table_name, sql_template, result_format, result_unit) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (metric_name, clean_name, "自动生成", "available", "L1",
                         table_name, sql_sum, "number", h),
                    )
                    count += 1
                except Exception as e:
                    logger.warning("指标生成(合计)失败: %s", e)

            # 平均
            metric_name = f"{clean_name} 平均"
            sql_avg = f'SELECT ROUND(AVG("{safe_h}"), 2) AS value FROM "{table_name}"'
            if not db.execute_one("SELECT 1 FROM metric_registry WHERE name=?", (metric_name,)):
                try:
                    db.execute_write(
                        "INSERT INTO metric_registry (name, display_name, category, status, complexity, table_name, sql_template, result_format, result_unit) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (metric_name, clean_name, "自动生成", "available", "L1",
                         table_name, sql_avg, "number", h),
                    )
                    count += 1
                except Exception as e:
                    logger.warning("指标生成(平均)失败: %s", e)

    return count


def _auto_register_kb_synonyms(headers: list, col_types: dict):
    """自动将列名注册为企业知识库同义词"""
    import yaml
    from pathlib import Path

    kb_path = Path(__file__).parent.parent.parent / "metrics" / "enterprise_kb.yaml"
    try:
        with open(kb_path, "r", encoding="utf-8") as f:
            kb = yaml.safe_load(f) or {}
    except Exception:
        kb = {}

    synonyms = kb.get("synonyms", {})
    if synonyms is None:
        synonyms = {}

    added = 0
    for h in headers:
        if col_types.get(h) in ("REAL", "INTEGER"):
            clean = _humanize_metric_name(h, "").strip()
            # 多种口语变体 → 指标名
            variants = [
                clean,
                h.replace('(', '').replace(')', ''),
                re.sub(r'\([^)]*\)', '', h).strip(),
            ]
            for v in variants:
                v2 = v.strip()
                if v2 and v2 not in synonyms:
                    # SUM variants
                    synonyms[v2] = f"{clean} 合计"
                    synonyms[f"{v2}合计"] = f"{clean} 合计"
                    synonyms[f"{v2}总和"] = f"{clean} 合计"
                    synonyms[f"{v2}总额"] = f"{clean} 合计"
                    # AVG variants
                    synonyms[f"{v2}平均"] = f"{clean} 平均"
                    synonyms[f"{v2}均值"] = f"{clean} 平均"
                    added += 1

    if added > 0:
        kb["synonyms"] = synonyms
        with open(kb_path, "w", encoding="utf-8") as f:
            yaml.dump(kb, f, allow_unicode=True, default_flow_style=False)
        logger.info("已注册 %d 个KB同义词到 %s", added, kb_path)
